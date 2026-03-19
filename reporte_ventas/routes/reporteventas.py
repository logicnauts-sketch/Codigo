from flask import Blueprint, render_template, jsonify, request, session, send_file
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from ..models.queries import (
    get_sales_kpis_db, get_top_productos_db, 
    get_ventas_por_hora_db, get_lista_ventas_db,
    get_alertas_db, get_tendencia_semanal_db,
    get_tendencia_mensual_db, get_analisis_dias_db
)

# Mock decorators for context
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('user_id'):
            return jsonify({"ok": False, "error": "Login required"}), 401
        return f(*args, **kwargs)
    return decorated_function

def solo_admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('rol') != 'admin':
            return jsonify({"ok": False, "error": "Admin required"}), 403
        return f(*args, **kwargs)
    return decorated_function

bp = Blueprint('reporteventas', __name__, template_folder='../templates', static_folder='../static')

@bp.route('/')
@login_required
@solo_admin_required
def index():
    nombre = session.get('nombre_completo', 'Administrador')
    rol = session.get('rol', 'admin')
    iniciales = "".join([n[0] for n in nombre.split()[:2] if n]).upper() if nombre else "AD"
    
    return render_template(
        "reporteventas.html", 
        now=datetime.now(),
        nombre_completo=nombre,
        rol=rol,
        iniciales=iniciales
    )

@bp.route('/api/kpis')
@login_required
@solo_admin_required
def api_kpis():
    period = request.args.get('period', 'hoy')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    result = get_sales_kpis_db(period, start_date, end_date)
    # Metas deshabilitadas hasta que se configuren desde el módulo
    result['metas'] = {"has_meta": False, "monto_objetivo": 0, "actual": result['totales']['ventas'], "porcentaje": 0, "proyeccion": 0}
    
    return jsonify(result)

@bp.route('/api/top-productos')
@login_required
@solo_admin_required
def api_top_productos():
    period = request.args.get('period', 'hoy')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    limit = int(request.args.get('limit', 10))
    productos = get_top_productos_db(period, start_date, end_date, limit)
    return jsonify(productos)

@bp.route('/api/ventas-hora')
@login_required
@solo_admin_required
def api_ventas_hora():
    period = request.args.get('period', 'hoy')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    datos = get_ventas_por_hora_db(period, start_date, end_date)
    return jsonify(datos)

@bp.route('/api/lista')
@login_required
@solo_admin_required
def api_lista():
    period = request.args.get('period', 'hoy')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    estado = request.args.get('estado')
    limit = int(request.args.get('limit', 200))
    ventas = get_lista_ventas_db(period, start_date, end_date, estado, limit)
    return jsonify(ventas)

@bp.route('/api/alertas')
@login_required
@solo_admin_required
def api_alertas():
    period = request.args.get('period', 'hoy')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    alertas = get_alertas_db(period, start_date, end_date)
    return jsonify(alertas)

@bp.route('/api/tendencia-semanal')
@login_required
@solo_admin_required
def api_tendencia_semanal():
    period = request.args.get('period', 'semana')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    tendencia = get_tendencia_semanal_db(period, start_date, end_date)
    return jsonify(tendencia)

@bp.route('/api/comparar', methods=['POST'])
@login_required
@solo_admin_required
def api_comparar():
    data = request.get_json()
    inicio_1 = data.get('inicio_1')
    fin_1 = data.get('fin_1')
    inicio_2 = data.get('inicio_2')
    fin_2 = data.get('fin_2')

    if not all([inicio_1, fin_1, inicio_2, fin_2]):
        return jsonify({"ok": False, "error": "Todas las fechas son requeridas"}), 400

    p1 = get_sales_kpis_db('personalizado', inicio_1, fin_1)
    p2 = get_sales_kpis_db('personalizado', inicio_2, fin_2)

    ventas_1 = p1.get('totales', {}).get('ventas', 0)
    ventas_2 = p2.get('totales', {}).get('ventas', 0)
    diferencia = ventas_1 - ventas_2
    pct = ((diferencia / ventas_2) * 100) if ventas_2 > 0 else 0

    return jsonify({
        "ok": True,
        "periodo_1": p1,
        "periodo_2": p2,
        "comparacion": {
            "diferencia_absoluta": diferencia,
            "diferencia_porcentaje": round(pct, 1)
        }
    })

@bp.route('/api/tendencia-mensual')
@login_required
@solo_admin_required
def api_tendencia_mensual():
    tendencia = get_tendencia_mensual_db()
    return jsonify(tendencia)

@bp.route('/api/analisis-dias')
@login_required
@solo_admin_required
def api_analisis_dias():
    period = request.args.get('period', 'semana')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    analisis = get_analisis_dias_db(period, start_date, end_date)
    return jsonify(analisis)

@bp.route('/api/export/excel')
@login_required
@solo_admin_required
def export_excel():
    period = request.args.get('period', 'hoy')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    ventas = get_lista_ventas_db(period, start_date, end_date, limit=5000)
    kpis = get_sales_kpis_db(period, start_date, end_date)
    
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    
    # Estilos
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Resumen KPI
    ws.append(["RESUMEN DEL PERIODO", period.upper()])
    ws.append(["Total Vendido", float(kpis['totales']['ventas'])])
    ws.append(["Cantidad Transacciones", kpis['totales']['cantidad']])
    ws.append(["Ticket Promedio", float(kpis['totales']['ticket_promedio'])])
    ws.append([]) # Espacio
    
    # Tabla de Ventas
    headers = ["ID", "FECHA", "CLIENTE", "METODO PAGO", "TOTAL", "ESTADO"]
    ws.append(headers)
    
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    for v in ventas:
        ws.append([
            v.get('id'),
            str(v.get('fecha_display')),
            v.get('cliente', 'S/C'),
            v.get('metodo_pago', 'Efectivo').upper(),
            float(v.get('total') or 0),
            v.get('estado', 'COMPLETADA').upper()
        ])
        
    wb.save(output)
    output.seek(0)
    filename = f"Reporte_Ventas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)

@bp.route('/api/export/pdf')
@login_required
@solo_admin_required
def export_pdf():
    period = request.args.get('period', 'hoy')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    ventas = get_lista_ventas_db(period, start_date, end_date, limit=1000)
    kpis = get_sales_kpis_db(period, start_date, end_date)
    
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    
    # Titulo
    elements.append(Paragraph("Next Design - REPORTE DE VENTAS", styles['Title']))
    elements.append(Paragraph(f"Periodo: {period.upper()} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # KPIs
    kpi_data = [
        ["Total Ventas", "Transacciones", "Ticket Promedio"],
        [f"RD$ {kpis['totales']['ventas']:,.2f}", str(kpis['totales']['cantidad']), f"RD$ {kpis['totales']['ticket_promedio']:,.2f}"]
    ]
    t_kpi = Table(kpi_data, colWidths=[180, 150, 180])
    t_kpi.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    elements.append(t_kpi)
    elements.append(Spacer(1, 24))
    
    # Tabla Detalle
    table_data = [["ID", "FECHA", "CLIENTE", "METODO", "TOTAL", "ESTADO"]]
    for v in ventas:
        table_data.append([
            str(v.get('id')),
            str(v.get('fecha_display'))[:16],
            (v.get('cliente') or 'S/C')[:20],
            v.get('metodo_pago', '').upper(),
            f"{float(v.get('total') or 0):,.2f}",
            v.get('estado', '').upper()
        ])
    
    t_ventas = Table(table_data, repeatRows=1)
    t_ventas.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'), # Montos a la derecha
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey)
    ]))
    elements.append(t_ventas)
    
    doc.build(elements)
    output.seek(0)
    filename = f"Reporte_Ventas_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(output, mimetype="application/pdf", as_attachment=True, download_name=filename)


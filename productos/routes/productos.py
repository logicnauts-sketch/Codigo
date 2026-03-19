from flask import render_template, session, jsonify, request, send_file
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from functools import wraps
from . . import bp

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            from flask import redirect, url_for
            return redirect(url_for('login.login'))
        return f(*args, **kwargs)
    return decorated

def solo_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('rol') != 'admin':
            return jsonify({"ok": False, "error": "Acceso denegado (Solo Admin)"}), 403
        return f(*args, **kwargs)
    return decorated

def sanitize_decimal(data):
    import decimal
    if isinstance(data, list):
        return [sanitize_decimal(v) for v in data]
    if isinstance(data, dict):
        return {k: sanitize_decimal(v) for k, v in data.items()}
    if isinstance(data, decimal.Decimal):
        return float(data)
    return data

@bp.route("/")
@login_required
def productos_page():
    return render_template("productos/productos.html", BRAND_NAME="Next Design")

@bp.route("/api/list", methods=["GET"])
@login_required
def api_list():
    from ..models.queries import get_productos_lista
    items = get_productos_lista()
    if items is None: items = []
    return jsonify({"ok": True, "items": sanitize_decimal(items)})

@bp.route("/api/next-code", methods=["GET"])
@login_required
def api_next_code():
    from ..models.queries import get_next_product_code
    code = get_next_product_code()
    return jsonify({"ok": True, "codigo": code})

@bp.route("/api/categorias", methods=["GET", "POST"])
@login_required
def api_categorias():
    from ..models.queries import get_categorias_lista, create_categoria
    if request.method == "POST":
        data = request.get_json()
        nombre = data.get("nombre")
        if not nombre: return jsonify({"ok": False, "error": "Nombre requerido"}), 400
        result = create_categoria(nombre)
        if result is None:
            return jsonify({"ok": False, "error": "No se pudo crear la categoría (podría ya existir)"}), 500
        return jsonify({"ok": True})
    
    cats = get_categorias_lista()
    if cats is None: cats = []
    return jsonify(sanitize_decimal(cats))

@bp.route("/api/categorias/<int:id>", methods=["DELETE"])
@login_required
@solo_admin_required
def api_categoria_delete(id):
    from ..models.queries import delete_categoria
    delete_categoria(id)
    return jsonify({"ok": True})

@bp.route("/api/create", methods=["POST"])
@login_required
@solo_admin_required
def api_create():
    from ..models.queries import create_producto
    data = request.get_json()
    try:
        pid = create_producto(data)
        return jsonify({"ok": True, "id": pid})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@bp.route("/api/update/<int:id>", methods=["POST"])
@login_required
@solo_admin_required
def api_update(id):
    from ..models.queries import update_producto
    data = request.get_json()
    try:
        update_producto(id, data)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@bp.route("/api/importar", methods=["POST"])
@login_required
@solo_admin_required
def api_importar():
    from ..models.queries import (
        get_producto_por_codigo, get_producto_por_barcode, 
        create_producto, update_producto, update_producto_estado
    )
    
    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "No hay archivo"}), 400
    
    file = request.files['file']
    try:
        wb = load_workbook(file)
        ws = wb.active
        rows = list(ws.rows)
        
        if len(rows) < 2:
            return jsonify({"ok": False, "error": "Archivo vacío"}), 400
        
        # Mapeo de cabeceras (Normalizar a minúsculas)
        headers = [str(cell.value).strip().upper() for cell in rows[0]]
        
        col_map = {
            'CODIGO': -1, 'BARCODE': -1, 'NOMBRE': -1, 
            'CATEGORÃA': -1, 'STOCK': -1, 'PRECIO VENTA': -1
        }
        
        for i, h in enumerate(headers):
            if 'CÃ“DIGO' in h or 'CODIGO' in h: col_map['CODIGO'] = i
            if 'BARCODE' in h or 'BARRA' in h: col_map['BARCODE'] = i
            if 'NOMBRE' in h: col_map['NOMBRE'] = i
            if 'CATEGORÃA' in h or 'CATEGORIA' in h: col_map['CATEGORÃA'] = i
            if 'STOCK' in h: col_map['STOCK'] = i
            if 'PRECIO' in h: col_map['PRECIO VENTA'] = i

        count_new = 0
        count_updated = 0
        
        def safe_val(val, as_str=True):
            if val is None: return "" if as_str else 0
            # Detectar errores comunes de Excel
            s_val = str(val).strip()
            if s_val.startswith('#'): return "" if as_str else 0
            
            if as_str:
                if isinstance(val, float) and val.is_integer():
                    return str(int(val))
                return s_val
            else:
                try: return float(val)
                except: return 0

        count_new = 0
        count_updated = 0
        
        for row_cells in rows[1:]:
            row = [cell.value for cell in row_cells]
            codigo = safe_val(row[col_map['CODIGO']]) if col_map['CODIGO'] != -1 else None
            barcode = safe_val(row[col_map['BARCODE']]) if col_map['BARCODE'] != -1 else None
            nombre = safe_val(row[col_map['NOMBRE']]) if col_map['NOMBRE'] != -1 else None
            
            if not nombre: continue 
            
            exists = None
            if codigo: exists = get_producto_por_codigo(codigo)
            if not exists and barcode: exists = get_producto_por_barcode(barcode)
            
            product_data = {
                'codigo': codigo,
                'codigo_barra': barcode,
                'nombre': nombre,
                'categoria': safe_val(row[col_map['CATEGORÃA']]) if col_map['CATEGORÃA'] != -1 else 'General',
                'precio_venta': safe_val(row[col_map['PRECIO VENTA']], False) if col_map['PRECIO VENTA'] != -1 else 0,
                'stock_actual': safe_val(row[col_map['STOCK']], False) if col_map['STOCK'] != -1 else 0,
                'stock_minimo': 5,
                'stock_maximo': 100,
                'descripcion': '',
                'precio_compra': 0,
                'impuesto': 18,
                'unidad_medida_id': None,
                'usa_dimension': 0
            }
            
            if exists:
                pid = exists['id']
                update_producto(pid, product_data)
                update_producto_estado(pid, 'Activo')
                count_updated += 1
            else:
                create_producto(product_data)
                count_new += 1
                
        return jsonify({
            "ok": True, 
            "msg": f"Importación exitosa. Nuevos: {count_new}, Actualizados/Reactivados: {count_updated}"
        })
        
    except Exception as e:
        print(f"[IMPORT ERROR] {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

@bp.route("/api/stats", methods=["GET"])
@login_required
def api_stats():
    from ..models.queries import get_stats_productos
    stats = get_stats_productos()
    if not stats: 
        stats = {"total_items": 0, "stock_bajo": 0, "valor_inventario": 0}
    return jsonify({"ok": True, **sanitize_decimal(stats)})

@bp.route("/api/profile/<int:id>", methods=["GET"])
@login_required
def api_profile(id):
    from ..models.queries import get_producto_por_id
    prod = get_producto_por_id(id)
    if not prod:
        return jsonify({"ok": False, "error": "Producto no encontrado"}), 404
    return jsonify({"ok": True, "producto": sanitize_decimal(prod), "historial_precios": []})

@bp.route("/api/toggle_status/<int:id>", methods=["POST"])
@login_required
@solo_admin_required
def api_toggle_status(id):
    from ..models.queries import get_producto_por_id, update_producto_estado
    prod = get_producto_por_id(id)
    if not prod:
        return jsonify({"ok": False, "error": "No existe"}), 404
    data = request.get_json(silent=True) or {}
    nuevo_estado = data.get('estado')
    
    if not nuevo_estado:
        nuevo_estado = "Inactivo" if prod['estado'] == 'Activo' else 'Activo'
        
    update_producto_estado(id, nuevo_estado)
    return jsonify({"ok": True, "nuevo_estado": nuevo_estado})

# Exportación a Excel Real
@bp.route('/api/exportar', methods=['GET'])
@login_required
@solo_admin_required
def exportar_productos():
    from ..models.queries import get_productos_lista
    items = get_productos_lista(include_inactive=True) or []
    
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalogo de Productos"
    headers = ["CÃ“DIGO", "BARCODE", "NOMBRE", "CATEGORÃA", "STOCK", "MINIMO", "PRECIO VENTA", "ESTADO"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for p in items:
        ws.append([
            p.get('codigo'), p.get('codigo_barra'), p.get('nombre'),
            p.get('categoria_nombre'), p.get('stock_actual'), p.get('stock_minimo'),
            float(p.get('precio_venta') or 0), p.get('estado')
        ])
        
    wb.save(output)
    output.seek(0)
    filename = f"Catalogo_Productos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)



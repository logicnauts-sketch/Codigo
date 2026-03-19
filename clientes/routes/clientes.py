from flask import Blueprint, render_template, session, jsonify, request, send_file
from functools import wraps
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from clientes.models.queries import (
    get_clientes_db, get_cliente_by_id, create_cliente_db, update_cliente_db,
    toggle_cliente_status, registrar_pago_db, get_credit_history_db,
    get_client_invoices_db, get_invoice_abonos_db, get_client_payments_db,
    get_client_by_document, crear_cliente_rapido_db, detect_schema
)
from infra.infra import execute_query

bp = Blueprint('clientes', __name__, 
               url_prefix='/clientes', 
               template_folder='../templates', 
               static_folder='../static')

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            # En el patrón Next Design, si no hay sesión, a veces forzamos una para desarrollo o redirigimos
            # Por ahora seguiremos el patrón de caja.js/app.py
            return jsonify({'success': False, 'message': 'Sesión requerida'}), 401
        return f(*args, **kwargs)
    return decorated_function

def solo_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('rol') != 'admin':
            return jsonify({'success': False, 'message': 'Acceso restringido a administradores'}), 403
        return f(*args, **kwargs)
    return decorated_function

@bp.route('/')
@login_required
def clientes():
    return render_template('clientes/clientes.html')

@bp.route('/api/clientes', methods=['GET'])
@login_required
def api_get_clientes():
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    try:
        clients, total = get_clientes_db(search, page, per_page)
        return jsonify({
            'success': True,
            'clients': clients,
            'total': total,
            'page': page,
            'per_page': per_page
        })
    except Exception as e:
        print(f"Error api_get_clientes: {e}")
        return jsonify({'success': False, 'message': 'Error al cargar clientes'}), 500

@bp.route('/api/clientes/<int:client_id>', methods=['GET'])
@login_required
def api_get_client(client_id):
    try:
        client = get_cliente_by_id(client_id)
        if not client:
            return jsonify({'success': False, 'message': 'Cliente no encontrado'}), 404
        return jsonify({'success': True, 'client': client})
    except Exception as e:
        print(f"Error api_get_client: {e}")
        return jsonify({'success': False, 'message': 'Error al cargar cliente'}), 500

@bp.route('/api/clientes', methods=['POST'])
@login_required
def api_create_client():
    data = request.get_json()
    if not data or not data.get('nombre'):
        return jsonify({'success': False, 'message': 'El nombre es obligatorio'}), 400
    
    try:
        new_id = create_cliente_db(data)
        client = get_cliente_by_id(new_id)
        return jsonify({'success': True, 'client': client})
    except Exception as e:
        print(f"Error api_create_client: {e}")
        return jsonify({'success': False, 'message': 'Error al crear el cliente'}), 500

@bp.route('/api/clientes/<int:client_id>', methods=['PUT'])
@login_required
def api_update_client(client_id):
    data = request.get_json()
    try:
        update_cliente_db(client_id, data)
        client = get_cliente_by_id(client_id)
        return jsonify({'success': True, 'client': client})
    except Exception as e:
        print(f"Error api_update_client: {e}")
        return jsonify({'success': False, 'message': 'Error al actualizar el cliente'}), 500

@bp.route('/api/clientes/<int:client_id>/toggle-status', methods=['PATCH'])
@login_required
@solo_admin_required
def api_toggle_client_status(client_id):
    try:
        new_status = toggle_cliente_status(client_id)
        if not new_status:
            return jsonify({'success': False, 'message': 'Cliente no encontrado'}), 404
        return jsonify({'success': True, 'new_status': new_status})
    except Exception as e:
        print(f"Error api_toggle_client_status: {e}")
        return jsonify({'success': False, 'message': 'Error al cambiar estado'}), 500

@bp.route('/api/clientes/<int:client_id>/pagar', methods=['POST'])
@login_required
def api_registrar_pago(client_id):
    data = request.get_json() or {}
    monto = data.get('monto_pago')
    metodo = data.get('metodo_pago')
    referencia = data.get('referencia_pago', '')
    credito_id = data.get('credito_id')

    if not monto or float(monto) <= 0:
        return jsonify({'success': False, 'message': 'Monto inválido'}), 400
    
    try:
        registrar_pago_db(
            client_id=client_id,
            monto_pago=monto,
            metodo=metodo,
            referencia=referencia,
            credito_id=credito_id,
            usuario_nombre=session.get('nombre_completo', 'Sistema')
        )
        return jsonify({'success': True, 'message': 'Pago registrado correctamente'})
    except Exception as e:
        print(f"Error api_registrar_pago: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/clientes/<int:client_id>/credito/historial', methods=['GET'])
@login_required
def api_get_credit_history(client_id):
    try:
        res = get_credit_history_db(client_id)
        if not res:
            return jsonify({'success': False, 'message': 'Cliente no encontrado'}), 404
        cliente, facturas, pagos = res
        return jsonify({
            'success': True,
            'cliente': cliente,
            'facturas_pendientes': facturas,
            'historial_pagos': pagos
        })
    except Exception as e:
        print(f"Error api_get_credit_history: {e}")
        return jsonify({'success': False, 'message': 'Error al cargar historial'}), 500

# Endpoint para KPIs básicos del reporte
@bp.route('/api/clientes/reporte/resumen', methods=['GET'])
@login_required
def api_reporte_clientes_resumen():
    try:
        total_cli = execute_query("SELECT COUNT(*) as total FROM clientes WHERE estado='activo'", fetch_one=True)['total']
        total_saldo = execute_query("SELECT SUM(saldo_actual) as total FROM clientes WHERE estado='activo'", fetch_one=True)['total'] or 0
        return jsonify({
            'success': True,
            'kpis': {
                'total_clientes': total_cli,
                'total_saldo_credito': float(total_saldo)
            }
        })
    except Exception as e:
        print(f"Error api_reporte_clientes_resumen: {e}")
        return jsonify({'success': False, 'message': 'Error al cargar resumen'}), 500

@bp.route('/api/clientes/<int:id>/facturas', methods=['GET'])
@login_required
def api_client_invoices(id):
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        only_pending = request.args.get('only_pending') == 'true'
        facturas = get_client_invoices_db(id, start_date, end_date, only_pending)
        return jsonify({'success': True, 'facturas': facturas})
    except Exception as e:
        print(f"Error api_client_invoices: {e}")
        return jsonify({'success': False, 'message': 'Error al cargar facturas'}), 500

@bp.route('/api/clientes/<int:id>/facturas/<int:factura_id>/abonos', methods=['GET'])
@login_required
def api_invoice_abonos(id, factura_id):
    try:
        resumen, abonos = get_invoice_abonos_db(id, factura_id)
        if not resumen:
            return jsonify({'success': False, 'message': 'No se encontraron abonos para esta factura'})
        return jsonify({
            'success': True,
            'resumen': resumen,
            'abonos': abonos
        })
    except Exception as e:
        print(f"Error api_invoice_abonos: {e}")
        return jsonify({'success': False, 'message': 'Error al cargar abonos'}), 500

@bp.route('/api/clientes/<int:id>/pagos', methods=['GET'])
@login_required
def api_client_payments(id):
    try:
        pagos = get_client_payments_db(id)
        return jsonify({'success': True, 'pagos': pagos})
    except Exception as e:
        print(f"Error api_client_payments: {e}")
        return jsonify({'success': False, 'message': 'Error al cargar pagos'}), 500

@bp.route('/api/clientes/buscar-por-documento', methods=['GET'])
@login_required
def api_buscar_por_documento():
    doc = request.args.get('documento')
    if not doc:
        return jsonify({'success': False, 'message': 'Documento requerido'}), 400
    try:
        client = get_client_by_document(doc)
        if client:
            return jsonify({'success': True, 'client': client})
        return jsonify({'success': False, 'message': 'Cliente no encontrado'}), 404
    except Exception as e:
        print(f"Error api_buscar_por_documento: {e}")
        return jsonify({'success': False, 'message': 'Error en la búsqueda'}), 500

@bp.route('/api/clientes/crear-rapido', methods=['POST'])
@login_required
def api_crear_rapido():
    try:
        data = request.get_json()
        nombre = data.get('nombre')
        rnc = data.get('rnc')
        if not nombre or not rnc:
            return jsonify({'success': False, 'message': 'Nombre y RNC son obligatorios'}), 400
            
        client_id = crear_cliente_rapido_db(nombre, rnc, data.get('telefono', ''), data.get('direccion', ''))
        client = get_cliente_by_id(client_id)
        return jsonify({
            'success': True, 
            'id': client_id, 
            'cliente': client,
            'message': 'Cliente creado correctamente'
        })
    except Exception as e:
        print(f"Error api_crear_rapido: {e}")
        return jsonify({'success': False, 'message': 'Error al crear cliente'}), 500

@bp.route('/api/clientes/validar-rnc/<rnc>', methods=['GET'])
@login_required
def api_validar_rnc(rnc):
    if not rnc:
        return jsonify({'success': False, 'message': 'RNC requerido'}), 400
    
    # 1. Buscar en BD local para ver si ya existe
    from ..models.queries import get_client_by_document
    cliente_existente = get_client_by_document(rnc)
    
    if cliente_existente:
        return jsonify({
            'success': True,
            'valido': True,
            'existe_en_bd': True,
            'cliente': {
                'id': cliente_existente['id'],
                'nombre': cliente_existente['nombre'],
                'cedula': cliente_existente.get('cedula', ''),
                'rnc': cliente_existente.get('rnc', '')
            },
            'message': 'Cliente ya registrado'
        })

    # 2. Simulación de validación DGII (en sistema real sería un API externo)
    is_valid = len(rnc) in (9, 11)
    
    if is_valid:
        tipo = 'RNC' if len(rnc) == 9 else 'CÃ‰DULA'
        return jsonify({
            'success': True, 
            'valido': True, 
            'existe_en_bd': False,
            'tipo': tipo,
            'message': f'{tipo} con formato válido',
            'data': {
                'nombre': f"ENTIDAD SIMULADA ({tipo})", # Mock de DGII
                'rnc_cedula': rnc
            }
        })
        
    return jsonify({
        'success': True, 
        'valido': False, 
        'existe_en_bd': False,
        'message': 'Formato de RNC/Cédula inválido'
    })

@bp.route('/api/clientes/reporte/excel', methods=['GET'])
@login_required
@solo_admin_required
def api_export_clientes_excel():
    try:
        status = request.args.get('status', 'activo')
        schema = detect_schema()
        email_col = schema['email_col']
        
        query = f"SELECT nombre, cedula as documento, telefono, {email_col} as email, direccion, estado, saldo_actual FROM clientes WHERE estado = %s"
        clients = execute_query(query, (status,))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Clientes"
        
        headers = ["Nombre", "Documento", "Teléfono", "Email", "Dirección", "Estado", "Deuda Actual"]
        ws.append(headers)
        
        for c in clients:
            ws.append([c['nombre'], c['documento'], c['telefono'], c['email'], c['direccion'], c['estado'], float(c['saldo_actual'] or 0)])
            
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(buffer, as_attachment=True, download_name=f"Reporte_Clientes_{datetime.now().strftime('%Y%m%d')}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        print(f"Error export excel: {e}")
        return jsonify({'success': False, 'message': 'Error al exportar Excel'}), 500

@bp.route('/api/clientes/reporte/pdf', methods=['GET'])
@login_required
@solo_admin_required
def api_export_clientes_pdf():
    # En esta versión simplificada, informamos que la exportación PDF requiere configuración adicional
    return jsonify({
        'success': False, 
        'message': 'La exportación a PDF (wkhtmltopdf) no está configurada en este entorno.',
        'info': 'Por favor use la exportación a Excel por el momento.'
    }), 200 # Devolvemos 200 para que el JS pueda manejar el mensaje amigablemente


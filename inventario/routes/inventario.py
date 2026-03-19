from flask import render_template, session, jsonify, request, send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from functools import wraps
from inventario import bp
from inventario.models.queries import (
    get_inventory_stats, get_inventario_lista, 
    get_historial_movimientos, registrar_movimiento_db,
    get_active_audit, iniciar_auditoria_db, get_audit_session_data,
    actualizar_conteo_db, finalizar_auditoria_db, cancelar_auditoria_db
)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({"ok": False, "error": "Sesión no iniciada"}), 401
        return f(*args, **kwargs)
    return decorated

@bp.route("/")
@login_required
def inventario_page():
    return render_template("inventario/inventario.html", BRAND_NAME="Next Design")

@bp.route('/api/stats')
@login_required
def api_stats():
    stats = get_inventory_stats()
    if stats:
        return jsonify({"ok": True, "stats": stats})
    return jsonify({"ok": False, "error": "No se pudieron obtener estadísticas"})

@bp.route('/api/list')
@login_required
def api_list():
    items = get_inventario_lista()
    return jsonify({"ok": True, "items": items})

@bp.route('/api/movimientos')
@login_required
def api_movimientos():
    limit = request.args.get('limit', 50, type=int)
    items = get_historial_movimientos(limit)
    return jsonify({"ok": True, "items": items})

@bp.route('/api/registrar-movimiento', methods=['POST'])
@login_required
def registrar_movimiento():
    data = request.json
    p_id = data.get('producto_id')
    tipo = data.get('tipo', 'Ajuste')
    cantidad = data.get('cantidad', 0)
    motivo = data.get('motivo', '')
    responsable = session.get('nombre_completo', 'Sistema')

    if not p_id or not cantidad:
        return jsonify({"ok": False, "error": "Producto y cantidad son requeridos"})

    try:
        nuevo_stock = registrar_movimiento_db(p_id, tipo, cantidad, responsable, motivo)
        return jsonify({
            "ok": True, 
            "msg": f"Movimiento de {tipo} registrado correctamente.",
            "nuevo_stock": nuevo_stock
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@bp.route('/api/audit/status')
@login_required
def api_audit_status():
    audit = get_active_audit()
    return jsonify({"ok": True, "active": bool(audit), "audit": audit})

@bp.route('/api/audit/iniciar', methods=['POST'])
@login_required
def api_audit_iniciar():
    try:
        if get_active_audit():
            return jsonify({"ok": False, "error": "Ya hay una sesión activa."})
        
        audit_id = iniciar_auditoria_db(session.get('user_id'))
        return jsonify({"ok": True, "msg": "Sesión de conteo iniciada.", "audit_id": audit_id})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@bp.route('/api/audit/session-data')
@login_required
def api_audit_data():
    audit = get_active_audit()
    if not audit:
        return jsonify({"ok": False, "error": "No hay sesión activa."})
    
    items = get_audit_session_data(audit['id'])
    # Convert Decimal to float for JSON compatibility
    for item in items:
        item['stock_teorico'] = float(item['stock_teorico'])
        item['stock_fisico'] = float(item['stock_fisico'])
        
    return jsonify({"ok": True, "items": items})

@bp.route('/api/audit/cargar-conteo', methods=['POST'])
@login_required
def api_audit_cargar_conteo():
    data = request.json
    d_id = data.get('detail_id')
    cantidad = data.get('cantidad', 0)
    
    if actualizar_conteo_db(d_id, cantidad):
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "No se pudo actualizar."})

@bp.route('/api/audit/finalizar', methods=['POST'])
@login_required
def api_audit_finalizar():
    audit = get_active_audit()
    if not audit:
        return jsonify({"ok": False, "error": "No hay sesión activa."})
    
    responsable = session.get('nombre_completo', 'Sistema')
    if finalizar_auditoria_db(audit['id'], responsable):
        return jsonify({"ok": True, "msg": "Conteo finalizado y stock actualizado."})
    return jsonify({"ok": False, "error": "No se pudo finalizar."})

@bp.route('/api/audit/cancelar', methods=['POST'])
@login_required
def api_audit_cancelar():
    audit = get_active_audit()
    if not audit:
        return jsonify({"ok": False, "error": "No hay sesión activa."})
    
    if cancelar_auditoria_db(audit['id']):
        return jsonify({"ok": True, "msg": "Sesión de conteo cancelada."})
    return jsonify({"ok": False, "error": "No se pudo cancelar."})

@bp.route('/api/exportar-excel')
@login_required
def exportar_excel():
    items = get_inventario_lista()
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Estilos
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["CÃ“DIGO", "PRODUCTO", "CATEGORÃA", "PRECIO", "COSTO", "STOCK"]
    ws.append(headers)
    
    for idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        ws.append([
            item['codigo'],
            item['nombre'],
            item['categoria'],
            item['precio'],
            item['costo'],
            item['stock']
        ])

    # Ajustar anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 5

    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="Inventario_Next Design.xlsx")

